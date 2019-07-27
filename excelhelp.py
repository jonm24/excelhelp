from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import urllib, json, tldextract, time, requests

# load excel sheet
wb = load_workbook(filename = 'AA_Directory_01.xlsx')
ws = wb['Sheet1']

# list to hold company names
companies = []

# get company names from excel sheet
for comps in ws.iter_cols(min_col=1, max_col=1, min_row=2, max_row=247, values_only=True):
	for i in comps:
		companies.append(i)

# initialize instance of webdriver
driver = webdriver.Chrome('/Users/jonmichalak/Desktop/chromedriver')

# global row variable
x = 2

# does process for every company in list
for comp in companies:
	# search for company website
	driver.get('https://www.google.com')
	elem = driver.find_element_by_name('q')
	elem.send_keys(comp + ' mi' + Keys.TAB + Keys.TAB + Keys.TAB + Keys.RETURN)

	# extract domain and write to excel
	url = driver.current_url
	ext = tldextract.extract(url)
	domain = '.'.join(ext[1:])
	ws.cell(row=x, column=4).value = domain

	# extract CMS and write to excel
	apiURL = "https://whatcms.org/APIEndpoint/Detect?key=2563dadee6918ef320e88a7e139587baff1f971ffffe70926ce7aa3b4060309f558f50&url=" + domain
	webdata = requests.get(apiURL)
	cmsdata = webdata.json()
	cms = cmsdata['result']['name']
	log = cmsdata['result']['code']	
	msg = cmsdata['result']['msg']
	ws.cell(row=x, column=2).value = cms
	ws.cell(row=x, column=5).value = log
	ws.cell(row=x, column=6).value = msg

	# save excel file, add to global row variable, get new window ready to repeat process
	wb.save(filename = 'AA_Directory_01.xlsx')
	x += 1
	driver.execute_script("window.open()")
	driver.switch_to.window(driver.window_handles[0])
	driver.close()
	driver.switch_to.window(driver.window_handles[0])
	time.sleep(11) 



